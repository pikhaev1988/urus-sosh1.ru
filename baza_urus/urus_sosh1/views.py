import openpyxl
from Tools.scripts.parse_html5_entities import create_dict
from django.core.serializers import serialize
from django.shortcuts import render
from django.urls import reverse_lazy, reverse
from urus_sosh1.models import *
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from urus_sosh1.form import register, MyImageForm
from django.http import HttpResponse
from django.shortcuts import render, redirect
import csv
import xlwt
import requests
from bs4 import BeautifulSoup
import mysql.connector
import json
from Tools.scripts.parse_html5_entities import create_dict

class regist(CreateView):
    form_class = register
    template_name = 'urus_sosh1/register/register.html'

def recvizit(req):
    return render(req, 'urus_sosh1/doc/recvixit.html')

def director(req):
    wer = doc.objects.all()
    return render(req, 'urus_sosh1/administraciy/direktor.html', {'wer': wer})

class Update(UpdateView):
    model = Pedagog
    template_name = 'urus_sosh1/Update.html'
    fields = '__all__'





def expor_csv(request, id):
    response = HttpResponse(content_type='text/xlsx')
    response['Content-Disposition'] = 'attachment; filename=export.csv'
    writer = csv.writer(response)
    venues = Pedagog.objects.filter(id=id)
    writer.writerow(['ФИО', ' Дата рождения', ' Адрес', ' СНИЛС', ' Телефон', ' Email'])
    for venue in venues:
        writer.writerow([venue.fio, venue.data_roj, venue.adres, venue.snils, venue.telefon, venue.email])
    return response


def export_excel(request, id):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="users.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Pedagog')

    row_num = 0

    font_styl =xlwt.XFStyle()
    font_styl.font.bold = True

    columns = ['ФИО', ' Дата рождения', ' Адрес', ' СНИЛС', ' Телефон', ' Email','Серия паспорта','Номер паспорта','Место работы','Стаж','Пед стаж','Категория','Должность','ВО СПО Учебное заведения','Спец по диплому','Сроки обучения','Серия диплома','Номер диплота','КПК (год прохождения последних КПК)','Место обучения и тема КПК']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_styl)

    font_styl = xlwt.XFStyle()

    rows = Pedagog.objects.filter(id=id).values_list('fio', 'data_roj', 'adres', 'snils', 'telefon', 'email','ser_pas','nom_pas','mesto_rab','staj','ped_staj','kate_gor','doljnost','uch_zav','spec_po_dip','sroc_ob','ser_dip','nom_dip','kpk_god_proh','mesto_ob')
    for row in rows:
        row_num +=1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_styl)
        wb.save(response)
        return response




def index(request):
    admin = doc.objects.all()
    categor = Category.objects.all()
    ped = Pedagog.objects.all()
    return render(request, "urus_sosh1/index.html", {'categor':categor, 'ped_1':ped,'admin':admin})

def pedagog(req, cet_id):
    ped = Pedagog.objects.filter(cet_id=cet_id)
    categor = Category.objects.all()
    cat = Category.objects.get(pk=cet_id)
    return render(req, 'urus_sosh1/pedagog.html', {'ped':ped, 'cat':cat, 'categor':categor})

class Detail(DetailView):
    model = Pedagog
    template_name = 'urus_sosh1/Detail.html'


def klass(request):
    wer = cat_uch.objects.all()
    return render(request, 'urus_sosh1/klass.html', {'wer':wer})

def uchenic(req, uch_id):
    uch = ucheniki.objects.all()
    qwe = ucheniki.objects.filter(uch_id=uch_id)
    kls = cat_uch.objects.filter(id=uch_id)
    rew = cat_uch.objects.all()
    return render(req, 'urus_sosh1/detail_klass.html', {'qwe':qwe, 'klas':kls, 'uch':uch, 'rew':rew})



class detal_uch(DetailView):
    model = ucheniki
    template_name = 'urus_sosh1/delail_uch.html'

def export_excel_uch(request, id):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="users.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('ucheniki')

    row_num = 0

    font_styl =xlwt.XFStyle()
    font_styl.font.bold = True

    columns = ['ФИО', ' Дата рождения', ' СНИЛС', 'Адрес ', ' ФИО Отца', ' Телефон', 'Место работы', ' ФИО Отца', ' Телефон', 'Место работы','Статус семьи']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_styl)

    font_styl = xlwt.XFStyle()

    rows = ucheniki.objects.filter(id=id).values_list('fio', 'data', 'snils', 'adres', 'fio_ot', 'tel_ot','mesto_rab_ot','fio_mat','tel_mat','mesto_rab_mat','status')
    for row in rows:
        row_num +=1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_styl)
        wb.save(response)
        return response


def zagruzca_failov(req):
    return render(req, 'urus_sosh1/zagruzka_faylov.html')




class trew(DetailView):
    model = ucheniki
    template_name = 'urus_sosh1/detail_klass.html'

def export_excel_klass(request, uch_id):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="users.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('ucheniki')

    row_num = 0

    font_styl =xlwt.XFStyle()
    font_styl.font.bold = True

    columns = ['ФИО', ' Дата рождения', ' СНИЛС', 'Адрес ', ' ФИО Отца', ' Телефон', 'Место работы', ' ФИО Отца', ' Телефон', 'Место работы','Статус семьи']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_styl)

    font_styl = xlwt.XFStyle()

    rows = ucheniki.objects.filter(id=uch_id).values_list('fio', 'data', 'snils', 'adres', 'fio_ot', 'tel_ot','mesto_rab_ot','fio_mat','tel_mat','mesto_rab_mat','status')
    ucen = ucheniki.objects.all()

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_styl)
    wb.save(response)
    return response, render(request,'urus_sosh1/detail_klass.html', {'uchen':ucen})




def donwlood(req):
    docs = doc.objects.filter(admini_id=4)
    return render(req, 'urus_sosh1/administraciy/direktor.html', {"doc":docs})

def ikt(req):
    docs = doc.objects.filter(admini_id=3)
    return render(req, 'urus_sosh1/administraciy/ikt.html', {"doc":docs})

def sekretar(req):
    docs = doc.objects.filter(admini_id=8)
    return render(req, 'urus_sosh1/administraciy/sekretar.html', {"doc":docs})

def psiholog(req):
    docs = doc.objects.filter(admini_id=9)
    return render(req, 'urus_sosh1/administraciy/psih.html', {"doc":docs})
def soc_ped(req):
    docs = doc.objects.filter(admini_id=7)
    return render(req, 'urus_sosh1/administraciy/soc_ped.html', {"doc":docs})
def zav_mr(req):
    docs = doc.objects.filter(admini_id=5)
    return render(req, 'urus_sosh1/administraciy/vaz_mr.html', {"doc":docs})
def zav_ahch(req):
    docs = doc.objects.filter(admini_id=6)
    return render(req, 'urus_sosh1/administraciy/zav_ahch.html', {"doc":docs})
def zav_vr(req):
    docs = doc.objects.filter(admini_id=2)
    return render(req, 'urus_sosh1/administraciy/zav_vr.html', {"doc":docs})
def zav_vr_nach(req):
    docs = doc.objects.filter(admini_id=1)
    return render(req, 'urus_sosh1/administraciy/zav_vr_nach.html', {"doc":docs})
def zav_yvr(req):
    docs = doc.objects.filter(admini_id=10)
    return render(req, 'urus_sosh1/administraciy/zav_yvr.html', {"doc":docs})


def download_data_ped(request, cet_id):
    queryset = Pedagog.objects.filter(cet_id=cet_id)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "MyModel data"
    worksheet.append(['id', 'name', 'value'])
    for obj in queryset:
        worksheet.append([obj.fio, obj.data_roj, obj.adres])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

    workbook.save(response)
    return response
#Доработать
def download_data_uch(request, uch_id):
    queryset = ucheniki.objects.filter(uch_id=uch_id)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "MyModel data"
    worksheet.append(['id', 'name', 'value'])
    for obj in queryset:
        worksheet.append([obj.fio, obj.data, obj.snils])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

    workbook.save(response)
    return response
#загрузка файлов
def upload_fail(request):
    if request.method == 'POST':
        form = MyImageForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            form=MyImageForm
            redirect(reverse('success'))  # redirect to a success page
    else:
        form = MyImageForm()
    return render(request, 'urus_sosh1/zagruzka_faylov.html', {'form': form})
